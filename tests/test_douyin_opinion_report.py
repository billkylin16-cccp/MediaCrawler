from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
import pytest

from store.douyin_opinion_report import DouyinOpinionReport
from tools.douyin_image_ocr import OcrPageResult


def _ts(hour: int) -> int:
    return int(datetime(2026, 8, 24, hour, tzinfo=timezone.utc).timestamp())


def test_windows_powershell_launcher_is_ascii_only():
    launcher = Path(__file__).parents[1] / "run_douyin_opinion.ps1"
    assert launcher.read_bytes().isascii()


def test_windows_double_click_launcher_is_portable():
    launcher = Path(__file__).parents[1] / "双击运行抖音舆情监测.cmd"
    content = launcher.read_bytes()
    assert content.isascii()
    assert b'%~dp0run_douyin_opinion.ps1' in content
    assert b'-PromptKeywords' in content


@pytest.mark.asyncio
async def test_report_filters_and_writes_requested_layout(tmp_path):
    output = tmp_path / "8.24抖音舆论检测.xlsx"
    report = DouyinOpinionReport(
        keywords=["武陟", "西陶"],
        target_date="2026-08-24",
        output_path=output,
        match="all",
    )
    assert report.add_video({
        "aweme_id": "video-1",
        "desc": "武陟西陶当天视频",
        "create_time": _ts(3),
        "author": {"nickname": "视频发布人"},
    })
    assert not report.add_video({
        "aweme_id": "video-2",
        "desc": "只有武陟",
        "create_time": _ts(3),
        "author": {"nickname": "不应收录"},
    })
    assert report.add_video({
        "aweme_id": "video-3",
        "desc": "=武陟西陶公式内容",
        "create_time": _ts(6),
        "author": {"nickname": "@公式发布人"},
    })
    await report.add_comments("video-1", [
        {
            "cid": "comment-1",
            "text": "今天的相关评论",
            "create_time": _ts(4),
            "user": {"nickname": "评论发布人"},
        },
        {
            "cid": "comment-2",
            "text": "非当天评论",
            "create_time": int(datetime(2026, 8, 25, 3, tzinfo=timezone.utc).timestamp()),
            "user": {"nickname": "不应收录"},
        },
        {
            "cid": "comment-3",
            "text": "=HYPERLINK(\"https://invalid.example\")",
            "create_time": _ts(5),
            "user": {"nickname": "+公式账号"},
        },
    ])
    report.flush()

    workbook = load_workbook(output, data_only=True)
    worksheet = workbook["舆论监测"]
    assert worksheet["A1"].value == "8.24抖音舆论检测"
    assert [worksheet.cell(2, index).value for index in range(1, 5)] == [
        "序号", "账号名称（发布人名称）", "发布内容", "关键信息"
    ]
    assert worksheet["B3"].value == "视频发布人"
    assert worksheet["B4"].value == "评论发布人"
    assert worksheet["B5"].value == "'+公式账号"
    assert worksheet["C5"].data_type != "f"
    assert worksheet["B6"].value == "'@公式发布人"
    assert worksheet["C6"].value == "'=武陟西陶公式内容"
    assert worksheet["C6"].data_type != "f"
    assert worksheet.max_row == 6
    workbook.close()


def test_report_any_match_and_invalid_output(tmp_path):
    report = DouyinOpinionReport(
        keywords=["WUZHI", "西陶"],
        target_date="2026-08-24",
        output_path=tmp_path / "report.xlsx",
        match="any",
    )
    assert report.add_video({
        "aweme_id": "video-any",
        "desc": "wuzhi 当天视频",
        "create_time": _ts(3),
        "author": {"nickname": "发布人"},
    })

    with pytest.raises(ValueError, match=".xlsx"):
        DouyinOpinionReport(
            keywords=["武陟"],
            target_date="2026-08-24",
            output_path=tmp_path / "report.csv",
        )

    existing = tmp_path / "existing.xlsx"
    existing.touch()
    with pytest.raises(FileExistsError, match="不会覆盖"):
        DouyinOpinionReport(
            keywords=["武陟"],
            target_date="2026-08-24",
            output_path=existing,
        )


def test_report_matches_carousel_ocr_and_marks_watch_account(tmp_path):
    output = tmp_path / "ocr-report.xlsx"
    report = DouyinOpinionReport(
        keywords=["西陶"],
        target_date="2026-08-24",
        output_path=output,
    )
    assert report.add_video(
        {
            "aweme_id": "image-post-1",
            "desc": "图片轮播",
            "create_time": _ts(3),
            "author": {"nickname": "重点发布人"},
        },
        ocr_pages=[
            OcrPageResult(page=1, text="无关内容", confidence=0.95),
            OcrPageResult(page=2, text="西陶镇相关信息", confidence=0.91),
        ],
        watch_account="重点发布人（xuhaoran888）",
    )
    report.flush()

    workbook = load_workbook(output, data_only=True)
    worksheet = workbook["舆论监测"]
    assert "【图片OCR】" in worksheet["C3"].value
    assert "第2张：西陶镇相关信息" in worksheet["C3"].value
    assert "命中来源：图片OCR第2张" in worksheet["D3"].value
    assert "重点账号：重点发布人（xuhaoran888）" in worksheet["D3"].value
    assert worksheet.row_dimensions[3].height > 54
    workbook.close()
