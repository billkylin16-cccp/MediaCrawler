# -*- coding: utf-8 -*-
# Copyright (c) 2025 relakkes@gmail.com
#
# This file is part of MediaCrawler project.
# Repository: https://github.com/NanmiCoder/MediaCrawler
# GitHub: https://github.com/NanmiCoder
# Licensed under NON-COMMERCIAL LEARNING LICENSE 1.1
#
# Local fork modification: compact, date-bounded Douyin review workbook.
# This module only processes data returned by the user's normal authenticated
# browser session. It does not perform network requests or bypass access limits.

"""Build the four-column Douyin opinion-monitor workbook.

The collector supplies raw public result objects to :class:`DouyinOpinionReport`.
This storage-layer module only filters them locally and writes the requested
review sheet; it does not make any network calls itself.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tools.douyin_image_ocr import OcrPageResult


CHINA_TZ = timezone(timedelta(hours=8))
HEADERS = ["序号", "账号名称（发布人名称）", "发布内容", "关键信息"]


@dataclass(frozen=True)
class VideoRecord:
    aweme_id: str
    account: str
    content: str
    created_at: datetime
    keywords: tuple[str, ...]
    description_keywords: tuple[str, ...]
    ocr_pages: tuple[OcrPageResult, ...]
    watch_account: str
    discovery_source: str

    @property
    def url(self) -> str:
        return f"https://www.douyin.com/video/{self.aweme_id}"


@dataclass(frozen=True)
class CommentRecord:
    comment_id: str
    aweme_id: str
    account: str
    content: str
    created_at: datetime
    keywords: tuple[str, ...]


def _china_today() -> date:
    return datetime.now(CHINA_TZ).date()


def _parse_target_date(value: str | date | None) -> date:
    if value is None or value == "":
        return _china_today()
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.date()
        return value.astimezone(CHINA_TZ).date()
    if isinstance(value, date):
        return value
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        raise ValueError("舆情日期必须为 YYYY-MM-DD，例如 2026-08-24")
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError("舆情日期必须为 YYYY-MM-DD，例如 2026-08-24") from exc


def _to_china_datetime(value: Any) -> Optional[datetime]:
    """Convert Douyin's seconds/milliseconds timestamp to China time."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=CHINA_TZ)
        return value.astimezone(CHINA_TZ)
    if isinstance(value, (int, float)) or str(value).strip().isdigit():
        timestamp = float(value)
        if timestamp > 20_000_000_000:
            timestamp /= 1000
        try:
            return datetime.fromtimestamp(timestamp, tz=timezone.utc).astimezone(CHINA_TZ)
        except (OSError, OverflowError, ValueError):
            return None
    text = str(value).strip().replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=CHINA_TZ)
    return parsed.astimezone(CHINA_TZ)


def _clean(value: Any) -> str:
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", str(value or ""))
    return " ".join(text.split())


def _excel_safe_text(value: Any) -> str:
    """Keep user-controlled text from being interpreted as an Excel formula."""
    text = _clean(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


class DouyinOpinionReport:
    """Accumulate selected videos/comments and export one review workbook."""

    def __init__(
        self,
        keywords: Iterable[str],
        target_date: str | date | None = None,
        output_path: str | Path | None = None,
        match: str = "all",
    ) -> None:
        self.keywords = tuple(item.strip() for item in keywords if item and item.strip())
        if not self.keywords:
            raise ValueError("舆情监测至少需要一个关键词")
        if match not in {"all", "any"}:
            raise ValueError("关键词匹配规则只能是 all 或 any")
        self.target_date = _parse_target_date(target_date)
        self.match = match
        default_name = f"{self.target_date.month}.{self.target_date.day:02d}抖音舆论检测.xlsx"
        self.output_path = Path(output_path) if output_path else Path(default_name)
        if self.output_path.suffix.lower() != ".xlsx":
            raise ValueError("舆情报表输出文件必须使用 .xlsx 扩展名")
        if self.output_path.exists():
            raise FileExistsError(
                f"舆情报表已存在，不会覆盖：{self.output_path}；请更换 --opinion_output"
            )
        self._videos: Dict[str, VideoRecord] = {}
        self._comments: Dict[str, CommentRecord] = {}

    def _matched_keywords(self, text: str) -> tuple[str, ...]:
        folded_text = text.casefold()
        return tuple(keyword for keyword in self.keywords if keyword.casefold() in folded_text)

    def _matches(self, hit: tuple[str, ...]) -> bool:
        return len(hit) == len(self.keywords) if self.match == "all" else bool(hit)

    def is_target_date(self, aweme: Dict[str, Any]) -> bool:
        return self.publish_date(aweme) == self.target_date

    @staticmethod
    def publish_date(aweme: Dict[str, Any]) -> Optional[date]:
        created_at = _to_china_datetime(aweme.get("create_time"))
        return created_at.date() if created_at else None

    def add_video(
        self,
        aweme: Dict[str, Any],
        ocr_pages: Iterable[OcrPageResult] = (),
        watch_account: str = "",
        discovery_source: str = "",
    ) -> bool:
        """Keep a video only when its text and publish date satisfy the report rule.

        The return value tells the caller whether comments for this video should
        be fetched. Duplicate search hits therefore do not trigger repeat work.
        """
        aweme_id = _clean(aweme.get("aweme_id"))
        if not aweme_id or aweme_id in self._videos:
            return False
        created_at = _to_china_datetime(aweme.get("create_time"))
        content = _clean(aweme.get("desc"))
        normalized_ocr_pages = tuple(
            OcrPageResult(
                page=page.page,
                text=_clean(page.text),
                confidence=page.confidence,
                source=page.source,
            )
            for page in ocr_pages
            if _clean(page.text)
        )
        description_hits = self._matched_keywords(content)
        combined_text = " ".join([content, *(page.text for page in normalized_ocr_pages)])
        hits = self._matched_keywords(combined_text)
        if not created_at or created_at.date() != self.target_date or not self._matches(hits):
            return False
        author = aweme.get("author") or {}
        account = _excel_safe_text(author.get("nickname")) or "未知账号"
        self._videos[aweme_id] = VideoRecord(
            aweme_id=aweme_id,
            account=account,
            content=_excel_safe_text(content) or "（视频未提供文字描述）",
            created_at=created_at,
            keywords=hits,
            description_keywords=description_hits,
            ocr_pages=normalized_ocr_pages,
            watch_account=_excel_safe_text(watch_account),
            discovery_source=_excel_safe_text(discovery_source),
        )
        return True

    async def add_comments(self, aweme_id: str, comments: List[Dict[str, Any]]) -> None:
        """Keep comments published on the monitored date for a selected video."""
        if aweme_id not in self._videos:
            return
        for comment in comments:
            comment_id = _clean(comment.get("cid"))
            comment_key = f"{aweme_id}:{comment_id}"
            if not comment_id or comment_key in self._comments:
                continue
            created_at = _to_china_datetime(comment.get("create_time"))
            if not created_at or created_at.date() != self.target_date:
                continue
            user = comment.get("user") or {}
            content = _clean(comment.get("text"))
            self._comments[comment_key] = CommentRecord(
                comment_id=comment_id,
                aweme_id=aweme_id,
                account=_excel_safe_text(user.get("nickname")) or "未知账号",
                content=_excel_safe_text(content) or "（评论未提供文字内容）",
                created_at=created_at,
                keywords=self._matched_keywords(content),
            )

    @staticmethod
    def _time_text(value: datetime) -> str:
        return value.strftime("%Y-%m-%d %H:%M")

    def _rows(self) -> list[list[Any]]:
        rows: list[list[Any]] = []
        videos = sorted(self._videos.values(), key=lambda item: (item.created_at, item.aweme_id))
        comments_by_video: Dict[str, list[CommentRecord]] = {}
        for comment in self._comments.values():
            comments_by_video.setdefault(comment.aweme_id, []).append(comment)

        for video in videos:
            matched = "、".join(video.keywords) or "搜索命中"
            source_parts: list[str] = []
            if video.description_keywords:
                source_parts.append("作品描述")
            ocr_hit_pages = [
                page.page
                for page in video.ocr_pages
                if page.source == "image" and self._matched_keywords(page.text)
            ]
            video_hit_frames = [
                page.page
                for page in video.ocr_pages
                if page.source == "video" and self._matched_keywords(page.text)
            ]
            if ocr_hit_pages:
                source_parts.append(
                    "图片OCR第" + "、".join(str(page) for page in ocr_hit_pages) + "张"
                )
            if video_hit_frames:
                source_parts.append(
                    "视频OCR第" + "、".join(str(page) for page in video_hit_frames) + "帧"
                )
            if not source_parts:
                source_parts.append("搜索命中")

            content_parts = [video.content]
            if video.ocr_pages:
                ocr_text = "｜".join(
                    (
                        f"第{page.page}{'帧' if page.source == 'video' else '张'}："
                        f"{page.text[:1000]}"
                    )
                    for page in video.ocr_pages
                )
                label = "视频抽帧OCR" if any(
                    page.source == "video" for page in video.ocr_pages
                ) else "图片OCR"
                content_parts.append(f"【{label}】{ocr_text[:12000]}")
            content = "\n".join(content_parts)

            info_parts = [
                "视频",
                f"发布时间：{self._time_text(video.created_at)}",
                f"匹配：{matched}",
                f"命中来源：{'、'.join(source_parts)}",
            ]
            if video.discovery_source:
                info_parts.append(f"发现方式：{video.discovery_source}")
            if video.watch_account:
                info_parts.append(f"重点账号：{video.watch_account}")
            info_parts.append(f"链接：{video.url}")
            rows.append([
                None,
                video.account,
                content,
                "｜".join(info_parts),
            ])
            for comment in sorted(comments_by_video.get(video.aweme_id, []), key=lambda item: (item.created_at, item.comment_id)):
                matched = "、".join(comment.keywords) or "关联视频"
                rows.append([
                    None,
                    comment.account,
                    f"【评论】{comment.content}",
                    f"评论｜发布时间：{self._time_text(comment.created_at)}｜关联视频发布人：{video.account}｜匹配：{matched}｜链接：{video.url}",
                ])
        for index, row in enumerate(rows, start=1):
            row[0] = index
        return rows

    def flush(self) -> Path:
        """Write the requested single-sheet Excel report, including an empty report."""
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "舆论监测"
        worksheet.sheet_view.showGridLines = False

        title = f"{self.target_date.month}.{self.target_date.day:02d}抖音舆论检测"
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = title
        worksheet["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14, color="FFFFFF")
        worksheet["A1"].fill = PatternFill("solid", fgColor="17365D")
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 28

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )
        for column, header in enumerate(HEADERS, start=1):
            cell = worksheet.cell(row=2, column=column, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        worksheet.row_dimensions[2].height = 26

        for row_number, values in enumerate(self._rows(), start=3):
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_number, column=column, value=value)
                cell.font = Font(name="Microsoft YaHei", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
            estimated_lines = max(
                4,
                (len(str(values[2] or "")) // 42) + 1,
                (len(str(values[3] or "")) // 48) + 1,
            )
            worksheet.row_dimensions[row_number].height = min(300, estimated_lines * 15)

        for column, width in {"A": 9, "B": 24, "C": 68, "D": 72}.items():
            worksheet.column_dimensions[column].width = width
        worksheet.freeze_panes = "A3"
        worksheet.auto_filter.ref = f"A2:D{max(2, worksheet.max_row)}"
        try:
            workbook.save(self.output_path)
        finally:
            workbook.close()
        return self.output_path

    @property
    def video_count(self) -> int:
        return len(self._videos)

    @property
    def comment_count(self) -> int:
        return len(self._comments)
